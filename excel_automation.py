# [file name]: excel_automation.py
# [file content begin]
import streamlit as st
import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import threading

class ExcelAutomation:
    """Class to handle automated Excel exports at regular intervals."""
    
    def __init__(self, save_folder, interval_minutes=3):
        self.save_folder = save_folder
        self.interval_minutes = interval_minutes
        self.is_running = False
        self.thread = None
        
        # Create save folder if it doesn't exist
        os.makedirs(save_folder, exist_ok=True)
    
    def start_automation(self, data_callback):
        """Start the automated Excel export process."""
        if self.is_running:
            st.warning("Automation is already running")
            return False
        
        self.is_running = True
        self.data_callback = data_callback
        
        # Start the automation in a separate thread
        self.thread = threading.Thread(target=self._automation_loop)
        self.thread.daemon = True
        self.thread.start()
        
        return True
    
    def stop_automation(self):
        """Stop the automated Excel export process."""
        self.is_running = False
        if self.thread:
            self.thread.join(timeout=1.0)
    
    def _automation_loop(self):
        """Main automation loop that runs at specified intervals."""
        while self.is_running:
            try:
                # Get data from callback function
                data = self.data_callback()
                if data:
                    # Generate Excel file
                    filename = self.generate_excel_report(data)
                    st.success(f"Automated report generated: {filename}")
                
                # Sleep for the specified interval
                time.sleep(self.interval_minutes * 60)
            except Exception as e:
                st.error(f"Error in automation loop: {e}")
                time.sleep(60)  # Wait a minute before retrying
    
    def generate_excel_report(self, data):
        """Generate a comprehensive Excel report with multiple sheets."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.save_folder, f"Market_Report_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add sheets for different data types
        if 'option_chain' in data and not data['option_chain'].empty:
            self._add_option_chain_sheet(wb, data['option_chain'], data.get('symbol', 'Unknown'))
        
        if 'analytics' in data:
            self._add_analytics_sheet(wb, data['analytics'])
        
        if 'technical_indicators' in data and not data['technical_indicators'].empty:
            self._add_technical_sheet(wb, data['technical_indicators'])
        
        if 'predictions' in data:
            self._add_predictions_sheet(wb, data['predictions'])
        
        if 'market_data' in data:
            self._add_market_data_sheet(wb, data['market_data'])
        
        # Save the workbook
        wb.save(filename)
        return filename
    
    def _add_option_chain_sheet(self, wb, df, symbol):
        """Add option chain data to Excel sheet."""
        ws = wb.create_sheet(f"{symbol} Option Chain")
        
        # Add headers
        headers = ['Strike', 'Call OI', 'Call Change', 'Call IV', 'Call LTP',
                  'Put LTP', 'Put IV', 'Put Change', 'Put OI']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_analytics_sheet(self, wb, analytics):
        """Add analytics data to Excel sheet."""
        ws = wb.create_sheet("Analytics")
        
        # Add headers
        ws.append(['Metric', 'Value'])
        
        # Add data
        for key, value in analytics.items():
            ws.append([key.replace('_', ' ').title(), value])
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_technical_sheet(self, wb, df):
        """Add technical indicators to Excel sheet."""
        ws = wb.create_sheet("Technical Indicators")
        
        # Add headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row in dataframe_to_rows(df, index=True, header=False):
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_predictions_sheet(self, wb, predictions):
        """Add predictions data to Excel sheet."""
        ws = wb.create_sheet("Predictions")
        
        # Add headers
        ws.append(['Model', 'Accuracy', 'Prediction', 'Confidence'])
        
        # Add data
        for model_name, result in predictions.items():
            prediction = "Bullish" if result.get('prediction', 0) == 1 else "Bearish"
            ws.append([
                model_name,
                result.get('accuracy', 0),
                prediction,
                result.get('confidence', 0)
            ])
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_market_data_sheet(self, wb, market_data):
        """Add market data to Excel sheet."""
        ws = wb.create_sheet("Market Data")
        
        # Add headers
        ws.append(['Index', 'Price', 'Change', 'Change %'])
        
        # Add data
        for index, data in market_data.items():
            if isinstance(data, dict) and 'price' in data:
                ws.append([
                    index,
                    data.get('price', 0),
                    data.get('change', 0),
                    data.get('change_pct', 0)
                ])
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

def create_excel_automation_tab():
    """Create the Excel automation tab."""
    st.markdown("### ⏰ Automated Excel Reports")
    
    st.info("""
    This feature automatically generates Excel reports every 3 minutes with the latest market data.
    Reports are saved to the 'excel_reports' folder with timestamps.
    """)
    
    # Initialize automation
    if 'excel_automation' not in st.session_state:
        st.session_state.excel_automation = ExcelAutomation('excel_reports', interval_minutes=3)
    
    # Control buttons
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🚀 Start Automation", use_container_width=True):
            # Define data callback function
            def get_data_callback():
                # This function should return the data to be exported
                # For now, we'll return mock data
                return {
                    'symbol': 'NIFTY',
                    'option_chain': pd.DataFrame({
                        'Strike': range(19500, 19650, 50),
                        'Call OI': np.random.randint(1000, 10000, 3),
                        'Call Change': np.random.randint(-1000, 1000, 3),
                        'Call IV': np.random.uniform(10, 30, 3),
                        'Call LTP': np.random.uniform(19500, 19800, 3),
                        'Put LTP': np.random.uniform(19500, 19800, 3),
                        'Put IV': np.random.uniform(10, 30, 3),
                        'Put Change': np.random.randint(-1000, 1000, 3),
                        'Put OI': np.random.randint(1000, 10000, 3)
                    }),
                    'analytics': {
                        'pcr': np.random.uniform(0.5, 1.5),
                        'max_pain': np.random.randint(19500, 19800),
                        'iv_percentile': np.random.uniform(20, 80),
                        'vix': np.random.uniform(15, 25)
                    },
                    'technical_indicators': pd.DataFrame({
                        'RSI': np.random.uniform(30, 70, 5),
                        'MACD': np.random.uniform(-2, 2, 5),
                        'Stochastic': np.random.uniform(20, 80, 5),
                        'ADX': np.random.uniform(20, 60, 5)
                    }, index=pd.date_range(end=datetime.now(), periods=5)),
                    'predictions': {
                        'Random Forest': {
                            'accuracy': np.random.uniform(0.7, 0.9),
                            'prediction': np.random.randint(0, 2),
                            'confidence': np.random.uniform(0.6, 0.95)
                        },
                        'XGBoost': {
                            'accuracy': np.random.uniform(0.7, 0.9),
                            'prediction': np.random.randint(0, 2),
                            'confidence': np.random.uniform(0.6, 0.95)
                        }
                    },
                    'market_data': {
                        'NIFTY': {
                            'price': np.random.uniform(19500, 19800),
                            'change': np.random.uniform(-100, 100),
                            'change_pct': np.random.uniform(-2, 2)
                        },
                        'BANKNIFTY': {
                            'price': np.random.uniform(43500, 44500),
                            'change': np.random.uniform(-200, 200),
                            'change_pct': np.random.uniform(-2, 2)
                        }
                    }
                }
            
            if st.session_state.excel_automation.start_automation(get_data_callback):
                st.success("Automation started! Reports will be generated every 3 minutes.")
    
    with col2:
        if st.button("⏹️ Stop Automation", use_container_width=True):
            st.session_state.excel_automation.stop_automation()
            st.info("Automation stopped.")
    
    # Display existing reports
    st.markdown("##### 📁 Generated Reports")
    
    if os.path.exists('excel_reports'):
        reports = [f for f in os.listdir('excel_reports') if f.endswith('.xlsx')]
        reports.sort(reverse=True)
        
        if reports:
            for report in reports[:5]:  # Show latest 5 reports
                report_path = os.path.join('excel_reports', report)
                report_time = os.path.getmtime(report_path)
                report_time_str = datetime.fromtimestamp(report_time).strftime("%Y-%m-%d %H:%M:%S")
                
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**{report}** (Generated: {report_time_str})")
                with col2:
                    with open(report_path, "rb") as file:
                        st.download_button(
                            label="Download",
                            data=file,
                            file_name=report,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{report}"
                        )
        else:
            st.info("No reports generated yet.")
    else:
        st.info("No reports folder found. Start automation to generate reports.")
    
    # Automation status
    st.markdown("##### 🔄 Automation Status")
    
    if st.session_state.excel_automation.is_running:
        st.success("✅ Automation is currently running")
        st.write(f"Reports are being generated every {st.session_state.excel_automation.interval_minutes} minutes")
    else:
        st.info("⏸️ Automation is not running")
    
    # Settings
    st.markdown("##### ⚙️ Automation Settings")
    
    new_interval = st.slider(
        "Report generation interval (minutes)",
        min_value=1,
        max_value=10,
        value=3,
        help="How often to generate Excel reports automatically"
    )
    
    if st.button("Update Interval"):
        st.session_state.excel_automation.interval_minutes = new_interval
        st.success(f"Interval updated to {new_interval} minutes")
# [file content end]