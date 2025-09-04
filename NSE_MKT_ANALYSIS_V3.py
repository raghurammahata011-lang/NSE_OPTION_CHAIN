# NSE Option Chain Analytics Pro - Enhanced Final Version
import numpy as np
import talib
from scipy.signal import argrelextrema
import time
import plotly.graph_objects as go
import random
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from sklearn.linear_model import LinearRegression, Ridge, Lasso, LogisticRegression
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, accuracy_score, confusion_matrix, classification_report
from sklearn.cluster import KMeans
import os
import concurrent.futures
import yfinance as yf
import warnings

# Import strategy functions
try:
    from strategies_final import option_greeks, plot_strategy, strategy_summary
except ImportError:
    try:
        from strategies_complete import option_greeks, plot_strategy, strategy_summary
    except ImportError:
        try:
            from strategies import option_greeks, plot_strategy, strategy_summary
        except ImportError:
            st.error("Please ensure strategies.py is in the same directory")

warnings.filterwarnings('ignore')

# ---------------- CONFIG ----------------
SAVE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "NSE_STOCK")
os.makedirs(SAVE_FOLDER, exist_ok=True)
THREAD_WORKERS = 6

# Initialize session state
if "indices" not in st.session_state:
    st.session_state.indices = ["NIFTY", "BANKNIFTY", "FINNIFTY"]
if "watchlist" not in st.session_state:
    st.session_state.watchlist = []
if "last_refresh" not in st.session_state:
    st.session_state.last_refresh = time.time()
if "auto_refresh" not in st.session_state:
    st.session_state.auto_refresh = False
if "custom_symbols" not in st.session_state:
    st.session_state.custom_symbols = []

# ---------------- ENHANCED SYMBOL MAPPING ----------------
import pandas as pd
import streamlit as st

# ----------------- MAJOR INDICES -----------------
COMPREHENSIVE_SYMBOL_MAP = {
    "NIFTY": "^NSEI",
    "BANKNIFTY": "^NSEBANK",
    "FINNIFTY": "NIFTY_FIN_SERVICE.NS",
    "MIDCPNIFTY": "NIFTY_MIDCAP_100.NS",
    "SENSEX": "^BSESN",
}

# ---------------- DYNAMIC SYMBOL LIST LOADING ----------------
try:
    # Load symbols from file
    intraday_df = pd.read_csv("symbol.txt")  # can also be .csv
    if 'SYMBOL' in intraday_df.columns:
        intraday_symbols_list = intraday_df['SYMBOL'].tolist()
    else:
        intraday_symbols_list = intraday_df.iloc[:, 0].tolist()

    st.success(f"Loaded {len(intraday_symbols_list)} intraday symbols from 'symbol.txt'.")

except FileNotFoundError:
    st.error("Error: 'symbol.txt' not found. Using fallback symbols.")
    intraday_symbols_list = ["NIFTY", "BANKNIFTY", "FINNIFTY", "RELIANCE"]
except Exception as e:
    st.error(f"An unexpected error occurred while loading 'symbol.txt': {e}")
    intraday_symbols_list = ["NIFTY", "BANKNIFTY", "FINNIFTY", "RELIANCE"]

# ---------------- DISPLAY SYMBOLS IN SIDEBAR ----------------
selected_symbol = st.sidebar.selectbox(
    "Select a symbol",
    options=intraday_symbols_list,
    index=0  # default selection
)

st.sidebar.write(f"You selected: {selected_symbol}")

# ---------------- PROFESSIONAL CSS ----------------
PRO_CSS = """
<style>
    .main-header {
        background: linear-gradient(90deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        padding: 2rem 1rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        text-align: center;
    }
    .metric-card {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #007acc;
        margin-bottom: 1rem;
    }
    .alert-box {
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        border-left: 4px solid;
    }
    .bullish { border-left-color: #28a745; background-color: #d4edda; }
    .bearish { border-left-color: #dc3545; background-color: #f8d7da; }
    .neutral { border-left-color: #6c757d; background-color: #e9ecef; }
    .stExpander > div:first-child {
        background-color: #f1f3f4;
    }
    .dataframe {
        font-size: 12px;
    }
</style>
"""

# ---------------- UTILITY: NSE Session ----------------
@st.cache_data(ttl=300)
def get_nse_session():
    """Create a requests session with common NSE headers to fetch option chain."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://www.nseindia.com/option-chain"
    })
    try:
        s.get("https://www.nseindia.com/option-chain", timeout=8)
    except Exception:
        pass
    return s

# ---------------- DATA TYPE UTILITIES ----------------
def ensure_dataframe_types(df):
    """Ensure all DataFrame columns have consistent data types for PyArrow compatibility."""
    if df.empty:
        return df
    
    # Convert object columns to appropriate types
    for col in df.columns:
        if df[col].dtype == 'object':
            # Try to convert to numeric first
            try:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except:
                # If that fails, convert to string
                df[col] = df[col].astype(str)
    
    return df

def create_safe_dataframe(data_dict):
    """Create a DataFrame with consistent data types."""
    df = pd.DataFrame(data_dict)
    
    # Ensure all values are strings for display purposes
    for col in df.columns:
        if col == 'Value':
            df[col] = df[col].astype(str)
    
    return df

# ---------------- ENHANCED SYMBOL RESOLUTION ----------------
def resolve_symbol_for_yfinance(symbol):
    """Enhanced symbol resolution for yfinance with multiple fallback options."""
    # Check if symbol is in our comprehensive mapping for indices
    if symbol.upper() in COMPREHENSIVE_SYMBOL_MAP:
        return COMPREHENSIVE_SYMBOL_MAP[symbol.upper()]

    # For all other stocks, use the symbol with the .NS suffix
    return f"{symbol.upper()}.NS"

def validate_symbol_exists(yf_symbol):
    """Validate if a yfinance symbol actually has data."""
    try:
        ticker = yf.Ticker(yf_symbol)
        hist = ticker.history(period="5d")
        return not hist.empty
    except:
        return False

def validate_symbol_exists(yf_symbol):
    """Validate if a yfinance symbol actually has data."""
    try:
        ticker = yf.Ticker(yf_symbol)
        hist = ticker.history(period="5d")
        return not hist.empty
    except:
        return False

# ---------------- NSE OPTION CHAIN SCRAPER (Enhanced) ----------------
def fetch_nse_options_alternative(symbol):
    """Enhanced alternative NSE option chain fetcher."""
    try:
        session = get_nse_session()
        
        # Build URL based on symbol type
        if symbol in st.session_state.indices:
            url = f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}"
        else:
            url = f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol}"
        
        # Multiple attempts with different headers
        for attempt in range(3):
            try:
                response = session.get(url, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                records = []
                for item in data.get("records", {}).get("data", []):
                    strike = item.get("strikePrice")
                    expiry = item.get("expiryDate")
                    
                    # CE data
                    ce = item.get("CE", {})
                    # PE data  
                    pe = item.get("PE", {})
                    
                    row = {
                        "Strike Price": strike,
                        "Expiry Date": expiry,
                        # CE
                        "CE_OI": ce.get("openInterest", 0),
                        "CE_Change_in_OI": ce.get("changeinOpenInterest", 0),
                        "CE_Volume": ce.get("totalTradedVolume", 0),
                        "CE_LTP": ce.get("lastPrice", 0),
                        "CE_IV": ce.get("impliedVolatility", 0),
                        # PE
                        "PE_OI": pe.get("openInterest", 0),
                        "PE_Change_in_OI": pe.get("changeinOpenInterest", 0),
                        "PE_Volume": pe.get("totalTradedVolume", 0),
                        "PE_LTP": pe.get("lastPrice", 0),
                        "PE_IV": pe.get("impliedVolatility", 0),
                    }
                    records.append(row)
                
                df = pd.DataFrame(records)
                return ensure_dataframe_types(df)
                
            except Exception as e:
                if attempt == 2:  # Last attempt
                    raise e
                time.sleep(random.uniform(1, 3))
        
    except Exception as e:
        st.error(f"Error fetching option chain for {symbol}: {e}")
        return pd.DataFrame()

# ---------------- FETCH / PARSE ----------------
def fetch_option_chain(symbol, session):
    """Enhanced option-chain fetcher with better error handling."""
    url = (f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}"
           if symbol in st.session_state.indices
           else f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol}")
    
    for attempt in range(3):
        try:
            resp = session.get(url, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            time.sleep(random.uniform(1.0, 2.5) * (attempt + 1))
            if attempt == 2:  # Last attempt
                st.warning(f"Failed to fetch option chain after 3 attempts: {str(e)[:100]}")
    return None

def parse_data(symbol, data):
    """Enhanced data parser with better error handling."""
    if not data:
        return pd.DataFrame()
    
    records = data.get("records", {}) or {}
    expiry_dates = records.get("expiryDates", []) or []
    
    if not expiry_dates:
        st.warning(f"No expiry dates found for {symbol}")
        return pd.DataFrame()
    
    # Get the nearest expiry
    expiry = expiry_dates[0]
    rows = []
    
    for item in records.get("data", []) or []:
        if item.get("expiryDate") != expiry:
            continue
        
        ce = item.get("CE") or {}
        pe = item.get("PE") or {}
        
        rows.append({
            "STRIKE": item.get("strikePrice"),
            "CALL_OI": ce.get("openInterest", 0),
            "CALL_CHNG_IN_OI": ce.get("changeinOpenInterest", 0),
            "CALL_IV": ce.get("impliedVolatility", 0),
            "CALL_LTP": ce.get("lastPrice", 0),
            "PUT_OI": pe.get("openInterest", 0),
            "PUT_CHNG_IN_OI": pe.get("changeinOpenInterest", 0),
            "PUT_IV": pe.get("impliedVolatility", 0),
            "PUT_LTP": pe.get("lastPrice", 0)
        })
    
    df = pd.DataFrame(rows)
    if df.empty:
        st.warning(f"No option data found for {symbol} for expiry {expiry}")
        return df
    
    # Ensure all numeric columns are properly typed
    numeric_columns = ['STRIKE', 'CALL_OI', 'CALL_CHNG_IN_OI', 'CALL_IV', 'CALL_LTP', 
                      'PUT_OI', 'PUT_CHNG_IN_OI', 'PUT_IV', 'PUT_LTP']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    return df.sort_values("STRIKE").reset_index(drop=True)

# ---------------- ENHANCED ANALYTICS ----------------
def calculate_analytics(df, spot_price=None):
    """Enhanced analytics with financial metrics."""
    if df.empty:
        return {}
    
    df = df.copy()
    
    # Estimate spot if not provided
    if not spot_price or spot_price == 0:
        if not df.empty and 'STRIKE' in df.columns and 'CALL_OI' in df.columns and 'PUT_OI' in df.columns:
            df['OI_DIFF'] = (df['CALL_OI'] - df['PUT_OI']).abs()
            if not df['OI_DIFF'].empty:
                spot_price = int(df.loc[df['OI_DIFF'].idxmin(), "STRIKE"])
            else:
                spot_price = int((df['STRIKE'].min() + df['STRIKE'].max()) / 2)
        else:
            spot_price = 0
    
    df['TOTAL_OI'] = df['CALL_OI'] + df['PUT_OI']
    df['OI_RATIO'] = df['PUT_OI'] / (df['CALL_OI'] + 1e-10)
    df['DELTA_CALL'] = df['CALL_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['DELTA_PUT'] = df['PUT_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['IV_DIFF'] = df['CALL_IV'] - df['PUT_IV']
    df['PRICE_RATIO'] = df['CALL_LTP'] / (df['PUT_LTP'] + 1e-10)
    
    total_call_oi = df['CALL_OI'].sum()
    total_put_oi = df['PUT_OI'].sum()
    pcr = total_put_oi / total_call_oi if total_call_oi > 0 else 0
    
    # ATM window
    strike_step = int(df['STRIKE'].diff().median() or 50)
    atm_strike = int(df.iloc[(df['STRIKE'] - spot_price).abs().argsort()[:1]]['STRIKE'].values[0])
    window = 3 * strike_step
    window_df = df[(df['STRIKE'] >= atm_strike - window) & (df['STRIKE'] <= atm_strike + window)]
    atm_pcr = window_df['PUT_OI'].sum() / window_df['CALL_OI'].sum() if window_df['CALL_OI'].sum() > 0 else 0
    
    # Max pain calculation
    strikes = df['STRIKE'].values
    losses = []
    for k in strikes:
        call_loss = ((np.clip(k - strikes, 0, None)) * df['CALL_OI']).sum()
        put_loss = ((np.clip(strikes - k, 0, None)) * df['PUT_OI']).sum()
        losses.append(call_loss + put_loss)
    max_pain = int(strikes[np.argmin(losses)]) if len(losses) > 0 else 0
    
    # Support/resistance by OI buildup
    df['PUT_SCORE'] = df['PUT_OI'] * (1 + df['PUT_CHNG_IN_OI'] / (df['PUT_OI'].replace(0, 1) + 1))
    df['CALL_SCORE'] = df['CALL_OI'] * (1 + df['CALL_CHNG_IN_OI'] / (df['CALL_OI'].replace(0, 1) + 1))
    
    strongest_support = int(df.loc[df['PUT_SCORE'].idxmax(), 'STRIKE'])
    strongest_resistance = int(df.loc[df['CALL_SCORE'].idxmax(), 'STRIKE'])
    
    # IV metrics
    avg_call_iv = df['CALL_IV'].mean()
    avg_put_iv = df['PUT_IV'].mean()
    iv_skew = round(avg_call_iv - avg_put_iv, 2)
    iv_atm = window_df[['CALL_IV', 'PUT_IV']].mean().mean() if not window_df.empty else (avg_call_iv + avg_put_iv) / 2
    
    # Expected move calculation
    expected_move_30d = round(spot_price * (iv_atm / 100) / np.sqrt(12), 2)
    
    # Directional score
    directional_score = (total_put_oi - total_call_oi) / (total_put_oi + total_call_oi) if (total_put_oi + total_call_oi) > 0 else 0
    directional_bias = "Bullish" if directional_score > 0.1 else ("Bearish" if directional_score < -0.1 else "Neutral")
    
    # Option flow analysis
    call_flow = (df['CALL_CHNG_IN_OI'] * df['CALL_LTP']).sum()
    put_flow = (df['PUT_CHNG_IN_OI'] * df['PUT_LTP']).sum()
    net_flow = call_flow - put_flow
    
    # Gamma exposure
    gamma_exposure = (df['CALL_OI'] + df['PUT_OI']).sum() / 1e6  # In millions
    
    # VIX-like indicator
    vix_like = (avg_call_iv + avg_put_iv) / 2
    
    # Create analytics summary without the full DataFrame
    analytics_summary = {
        "pcr": round(pcr, 2),
        "atm_pcr": round(atm_pcr, 2),
        "max_pain": max_pain,
        "support": strongest_support,
        "resistance": strongest_resistance,
        "avg_call_iv": round(avg_call_iv, 2),
        "avg_put_iv": round(avg_put_iv, 2),
        "iv_skew": iv_skew,
        "iv_atm": round(iv_atm, 2),
        "expected_move_30d": expected_move_30d,
        "directional_score": round(directional_score, 2),
        "directional_bias": directional_bias,
        "call_flow": round(call_flow, 2),
        "put_flow": round(put_flow, 2),
        "net_flow": round(net_flow, 2),
        "gamma_exposure": round(gamma_exposure, 2),
        "vix_like": round(vix_like, 2),
        "spot_price": spot_price
    }
    
    return analytics_summary

# ---------------- ENHANCED FINANCIAL ANALYTICS ----------------
def fetch_historical_data(symbol, period="6mo"):
    """Enhanced historical data fetching with comprehensive symbol resolution."""
    try:
        # Get the correct yfinance symbol  
        yf_symbol = resolve_symbol_for_yfinance(symbol)
        
        # Validate symbol exists
        if not validate_symbol_exists(yf_symbol):
            st.warning(f"Symbol {yf_symbol} may not have recent data")
        
        ticker = yf.Ticker(yf_symbol)
        hist = ticker.history(period=period)
        
        # If we still don't get data, try alternative periods
        if hist.empty:
            alternative_periods = ["3mo", "1mo", "5d", "1d"]
            for alt_period in alternative_periods:
                hist = ticker.history(period=alt_period)
                if not hist.empty:
                    st.info(f"Using {alt_period} data instead of {period}")
                    break
        
        # Get additional info for context
        if not hist.empty:
            try:
                info = ticker.info
                st.session_state[f"{symbol}_info"] = {
                    'longName': info.get('longName', symbol),
                    'sector': info.get('sector', 'Unknown'),
                    'marketCap': info.get('marketCap', 0),
                    'currency': info.get('currency', 'INR')
                }
            except:
                pass
        
        return hist
        
    except Exception as e:
        st.error(f"Error fetching historical data for {symbol}: {e}")
        return pd.DataFrame()

def calculate_technical_indicators(df):
    """Enhanced technical indicators with smooth calculations."""
    if df.empty:
        return df
    
    # Calculate simple moving averages with smooth interpolation
    df['SMA_5'] = df['Close'].rolling(window=5, min_periods=1).mean()
    df['SMA_10'] = df['Close'].rolling(window=10, min_periods=1).mean()
    df['SMA_20'] = df['Close'].rolling(window=20, min_periods=1).mean()
    df['SMA_50'] = df['Close'].rolling(window=50, min_periods=1).mean()
    
    # Exponential moving averages for smoother curves
    df['EMA_12'] = df['Close'].ewm(span=12, adjust=False).mean()
    df['EMA_26'] = df['Close'].ewm(span=26, adjust=False).mean()
    
    # Calculate RSI with smooth handling
    delta = df['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14, min_periods=1).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14, min_periods=1).mean()
    rs = gain / (loss + 1e-10)  # Avoid division by zero
    df['RSI'] = 100 - (100 / (1 + rs))
    
    # Calculate MACD with smooth signals
    df['MACD'] = df['EMA_12'] - df['EMA_26']
    df['MACD_Signal'] = df['MACD'].ewm(span=9, adjust=False).mean()
    df['MACD_Histogram'] = df['MACD'] - df['MACD_Signal']
    
    # Calculate Bollinger Bands
    df['BB_Middle'] = df['Close'].rolling(window=20, min_periods=1).mean()
    bb_std = df['Close'].rolling(window=20, min_periods=1).std()
    df['BB_Upper'] = df['BB_Middle'] + (bb_std * 2)
    df['BB_Lower'] = df['BB_Middle'] - (bb_std * 2)
    
    # Additional indicators
    # Stochastic Oscillator
    low_14 = df['Low'].rolling(window=14, min_periods=1).min()
    high_14 = df['High'].rolling(window=14, min_periods=1).max()
    df['%K'] = 100 * ((df['Close'] - low_14) / (high_14 - low_14 + 1e-10))
    df['%D'] = df['%K'].rolling(window=3, min_periods=1).mean()
    
    # Williams %R
    df['Williams_R'] = -100 * ((high_14 - df['Close']) / (high_14 - low_14 + 1e-10))
    
    # Average True Range
    high_low = df['High'] - df['Low']
    high_close = np.abs(df['High'] - df['Close'].shift())
    low_close = np.abs(df['Low'] - df['Close'].shift())
    true_ranges = pd.concat([high_low, high_close, low_close], axis=1)
    df['ATR'] = true_ranges.max(axis=1).rolling(window=14, min_periods=1).mean()
    
    return df

# ---------------- ENHANCED ML PREDICTIONS ----------------
def predict_price_movement_enhanced(hist_data, analytics, df_options=None):
    """Enhanced ML prediction with better feature engineering."""
    try:
        if hist_data.empty:
            return create_fallback_predictions(analytics, df_options)
        
        # Ensure we have sufficient data
        if len(hist_data) < 10:
            return create_fallback_predictions(analytics, df_options)
        
        # Prepare enhanced features from historical data
        features = hist_data[['Open', 'High', 'Low', 'Close', 'Volume']].copy()
        
        # Add technical indicators as features
        features['Returns'] = features['Close'].pct_change()
        features['Log_Returns'] = np.log(features['Close'] / features['Close'].shift(1))
        features['Volatility'] = features['Returns'].rolling(10, min_periods=1).std()
        features['Price_Range'] = (features['High'] - features['Low']) / features['Close']
        features['Volume_MA'] = features['Volume'].rolling(10, min_periods=1).mean()
        features['Volume_Ratio'] = features['Volume'] / features['Volume_MA']
        
        # Add moving average ratios
        if 'SMA_20' in hist_data.columns:
            features['SMA_Ratio_20'] = features['Close'] / hist_data['SMA_20']
        if 'SMA_50' in hist_data.columns:
            features['SMA_Ratio_50'] = features['Close'] / hist_data['SMA_50']
        if 'RSI' in hist_data.columns:
            features['RSI'] = hist_data['RSI']
        if 'MACD' in hist_data.columns:
            features['MACD'] = hist_data['MACD']
        
        # Add option-based features if available
        if analytics:
            features['PCR'] = analytics.get('pcr', 1)
            features['IV_ATM'] = analytics.get('iv_atm', 20)
            features['Max_Pain_Distance'] = abs(features['Close'].iloc[-1] - analytics.get('max_pain', features['Close'].iloc[-1]))
            features['IV_Skew'] = analytics.get('iv_skew', 0)
            features['Net_Flow'] = analytics.get('net_flow', 0)
            features['Gamma_Exposure'] = analytics.get('gamma_exposure', 0)
        
        # Target: Next day return direction (1 for up, 0 for down)
        features['Target'] = (features['Close'].shift(-1) > features['Close']).astype(int)
        
        # Drop NaN values
        features = features.dropna()
        
        if len(features) < 5:
            return create_fallback_predictions(analytics, df_options)
        
        # Prepare data for ML
        X = features.drop(['Target'], axis=1)
        y = features['Target']
        
        # Handle case where we don't have enough data for splitting
        if len(features) < 10:
            # Use all data for training and testing
            X_train = X_test = X
            y_train = y_test = y
        else:
            # Split data (use more data for training if available)
            test_size = min(0.3, max(0.1, 5/len(features)))
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)
        
        # Scale features
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)
        
        # Train multiple models with reduced complexity for small datasets
        models = {
            'Random Forest': RandomForestClassifier(n_estimators=min(50, len(features)*2), 
                                                   random_state=42, max_depth=3),
            'Logistic Regression': LogisticRegression(random_state=42, max_iter=1000, 
                                                     solver='liblinear'),
        }
        
        # Add Gradient Boosting only if we have enough data
        if len(features) >= 15:
            models['Gradient Boosting'] = GradientBoostingRegressor(n_estimators=30, 
                                                                   random_state=42, max_depth=2)
        
        results = {}
        for name, model in models.items():
            try:
                if name == 'Gradient Boosting':
                    model.fit(X_train_scaled, y_train.astype(float))
                    pred_proba = model.predict(X_test_scaled)
                    pred = (pred_proba > 0.5).astype(int)
                    # Get latest prediction
                    latest_pred = (model.predict(X_train_scaled[-1:]) > 0.5).astype(int)[0]
                else:
                    model.fit(X_train_scaled, y_train)
                    pred = model.predict(X_test_scaled)
                    if hasattr(model, 'predict_proba'):
                        pred_proba = model.predict_proba(X_test_scaled)[:, 1]
                        latest_pred = model.predict(X_train_scaled[-1:])[0]
                    else:
                        latest_pred = pred[-1] if len(pred) > 0 else 0
                
                accuracy = accuracy_score(y_test, pred) if len(y_test) > 0 else 0.5
                
                results[name] = {
                    'accuracy': accuracy,
                    'model': model,
                    'prediction': int(latest_pred),
                    'confidence': abs(accuracy - 0.5) * 2  # Convert to confidence score
                }
                
            except Exception as e:
                print(f"Error training {name}: {e}")
                continue
        
        # Add ensemble prediction if we have multiple models
        if len(results) > 1:
            predictions = [r['prediction'] for r in results.values()]
            accuracies = [r['accuracy'] for r in results.values()]
            
            # Weighted ensemble
            weights = np.array(accuracies) / sum(accuracies) if sum(accuracies) > 0 else np.ones(len(accuracies)) / len(accuracies)
            ensemble_pred = 1 if np.average(predictions, weights=weights) > 0.5 else 0
            ensemble_accuracy = np.average(accuracies)
            
            results['Ensemble'] = {
                'accuracy': ensemble_accuracy,
                'prediction': ensemble_pred,
                'confidence': max([r['confidence'] for r in results.values()]),
                'model': 'Weighted Ensemble'
            }
        
        return results
        
    except Exception as e:
        st.error(f"ML Prediction Error: {e}")
        return create_fallback_predictions(analytics, df_options)

def create_fallback_predictions(analytics, df_options=None):
    """Enhanced fallback predictions based on option chain analytics."""
    try:
        predictions = {}
        
        # PCR-based prediction
        pcr = analytics.get('pcr', 1) if analytics else 1
        pcr_prediction = 1 if pcr > 1.2 else 0
        pcr_confidence = min(abs(pcr - 1) * 0.5, 0.8)
        
        predictions['PCR Model'] = {
            'accuracy': 0.55 + pcr_confidence * 0.2,
            'prediction': pcr_prediction,
            'confidence': pcr_confidence
        }
        
        # IV Skew-based prediction
        if analytics:
            iv_skew = analytics.get('iv_skew', 0)
            skew_prediction = 1 if iv_skew < 0 else 0  # Negative skew often bullish
            skew_confidence = min(abs(iv_skew) * 0.1, 0.7)
            
            predictions['IV Skew Model'] = {
                'accuracy': 0.53 + skew_confidence * 0.15,
                'prediction': skew_prediction,
                'confidence': skew_confidence
            }
        
        # Max Pain-based prediction
        if analytics:
            spot_price = analytics.get('spot_price', 0)
            max_pain = analytics.get('max_pain', 0)
            if spot_price > 0 and max_pain > 0:
                pain_diff = (spot_price - max_pain) / spot_price
                pain_prediction = 1 if pain_diff < -0.02 else (0 if pain_diff > 0.02 else 1)
                pain_confidence = min(abs(pain_diff) * 5, 0.6)
                
                predictions['Max Pain Model'] = {
                    'accuracy': 0.52 + pain_confidence * 0.18,
                    'prediction': pain_prediction,
                    'confidence': pain_confidence
                }
        
        return predictions
        
    except Exception as e:
        print(f"Error creating fallback predictions: {e}")
        return {
            'Basic Model': {
                'accuracy': 0.5,
                'prediction': 1,
                'confidence': 0.1
            }
        }

# ---------------- ENHANCED SMOOTH VISUALIZATIONS ----------------
def plot_oi_distribution(df, spot_price):
    """Enhanced OI distribution chart with smooth styling."""
    if df.empty:
        return go.Figure()
    
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        subplot_titles=('Open Interest Distribution', 'OI Change'),
                        vertical_spacing=0.08)
    
    # OI Distribution with enhanced styling
    fig.add_trace(go.Bar(
        x=df['STRIKE'], 
        y=df['CALL_OI'], 
        name='Call OI', 
        marker=dict(color='rgba(239, 85, 59, 0.8)', line=dict(color='rgb(239, 85, 59)', width=1)),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Call OI:</b> %{y:,.0f}<extra></extra>'
    ), row=1, col=1)
    
    fig.add_trace(go.Bar(
        x=df['STRIKE'], 
        y=-df['PUT_OI'], 
        name='Put OI',
        marker=dict(color='rgba(46, 204, 113, 0.8)', line=dict(color='rgb(46, 204, 113)', width=1)),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Put OI:</b> %{y:,.0f}<extra></extra>'
    ), row=1, col=1)
    
    # OI Change with enhanced styling
    fig.add_trace(go.Bar(
        x=df['STRIKE'], 
        y=df['CALL_CHNG_IN_OI'], 
        name='Call OI Change',
        marker=dict(color='rgba(192, 57, 43, 0.8)', line=dict(color='rgb(192, 57, 43)', width=1)),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Call OI Change:</b> %{y:,.0f}<extra></extra>'
    ), row=2, col=1)
    
    fig.add_trace(go.Bar(
        x=df['STRIKE'], 
        y=-df['PUT_CHNG_IN_OI'], 
        name='Put OI Change',
        marker=dict(color='rgba(39, 174, 96, 0.8)', line=dict(color='rgb(39, 174, 96)', width=1)),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Put OI Change:</b> %{y:,.0f}<extra></extra>'
    ), row=2, col=1)
    
    # Add spot price line with enhanced styling
    fig.add_vline(
        x=spot_price, 
        line=dict(width=3, dash="dash", color="rgba(52, 152, 219, 0.8)"), 
        annotation=dict(text="Current Spot", textangle=90, font=dict(size=12, color="rgb(52, 152, 219)"))
    )
    
    fig.update_layout(
        height=650,
        showlegend=True,
        title_font_size=16,
        plot_bgcolor='rgba(248, 249, 250, 0.8)',
        paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=11),
        hovermode='x unified'
    )
    
    fig.update_xaxes(title_text="Strike Price", row=2, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(title_text="Open Interest", row=1, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(title_text="OI Change", row=2, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    
    return fig

def plot_oi_change(df, spot_price):
    """Enhanced OI change analysis with smooth lines."""
    if df.empty:
        return go.Figure()
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['CALL_CHNG_IN_OI'], 
        mode='lines+markers', 
        name='Call OI Change',
        line=dict(color='rgb(231, 76, 60)', width=3, smoothing=1.3),
        marker=dict(size=6, color='rgb(231, 76, 60)', line=dict(width=1, color='white')),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Call OI Change:</b> %{y:,.0f}<extra></extra>'
    ))
    
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['PUT_CHNG_IN_OI'],
        mode='lines+markers', 
        name='Put OI Change', 
        line=dict(color='rgb(46, 204, 113)', width=3, smoothing=1.3),
        marker=dict(size=6, color='rgb(46, 204, 113)', line=dict(width=1, color='white')),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Put OI Change:</b> %{y:,.0f}<extra></extra>'
    ))
    
    fig.add_vline(
        x=spot_price, 
        line=dict(width=3, dash="dash", color="rgba(52, 152, 219, 0.8)"),
        annotation=dict(text="Spot Price", textangle=90, font=dict(size=12, color="rgb(52, 152, 219)"))
    )
    
    fig.update_layout(
        title="Open Interest Change Analysis",
        xaxis_title="Strike Price",
        yaxis_title="Change in OI",
        height=450,
        showlegend=True,
        plot_bgcolor='rgba(248, 249, 250, 0.8)',
        paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=11),
        hovermode='x unified'
    )
    
    fig.update_xaxes(gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(gridcolor='rgba(128, 128, 128, 0.2)')
    
    return fig

def plot_iv_surface(df, spot_price):
    """Enhanced IV surface/skew with smooth curves."""
    if df.empty:
        return go.Figure()
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['CALL_IV'],
        mode='lines+markers', 
        name='Call IV',
        line=dict(color='rgb(155, 89, 182)', width=3, smoothing=1.3),
        marker=dict(size=6, color='rgb(155, 89, 182)', line=dict(width=1, color='white')),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Call IV:</b> %{y:.2f}%<extra></extra>'
    ))
    
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['PUT_IV'],
        mode='lines+markers', 
        name='Put IV',
        line=dict(color='rgb(241, 196, 15)', width=3, smoothing=1.3),
        marker=dict(size=6, color='rgb(241, 196, 15)', line=dict(width=1, color='white')),
        hovertemplate='<b>Strike:</b> %{x}<br><b>Put IV:</b> %{y:.2f}%<extra></extra>'
    ))
    
    fig.add_vline(
        x=spot_price, 
        line=dict(width=3, dash="dash", color="rgba(52, 152, 219, 0.8)"),
        annotation=dict(text="Spot Price", textangle=90, font=dict(size=12, color="rgb(52, 152, 219)"))
    )
    
    fig.update_layout(
        title="Implied Volatility Surface",
        xaxis_title="Strike Price", 
        yaxis_title="Implied Volatility (%)",
        height=450,
        showlegend=True,
        plot_bgcolor='rgba(248, 249, 250, 0.8)',
        paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=11),
        hovermode='x unified'
    )
    
    fig.update_xaxes(gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(gridcolor='rgba(128, 128, 128, 0.2)')
    
    return fig

def plot_technical_indicators(hist_data):
    """Enhanced technical indicators with smooth styling and more indicators."""
    if hist_data.empty:
        return go.Figure()
    
    fig = make_subplots(
        rows=4, cols=1, 
        shared_xaxes=True,
        subplot_titles=('Price & Moving Averages', 'RSI & Stochastic', 'MACD', 'Volume & ATR'),
        vertical_spacing=0.06,
        row_heights=[0.4, 0.2, 0.2, 0.2]
    )
    
    # Price and Moving Averages
    fig.add_trace(go.Scatter(
        x=hist_data.index, 
        y=hist_data['Close'],
        name='Close Price', 
        line=dict(color='rgb(52, 152, 219)', width=2),
        hovertemplate='<b>Date:</b> %{x}<br><b>Close:</b> ₹%{y:.2f}<extra></extra>'
    ), row=1, col=1)
    
    if 'SMA_20' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['SMA_20'],
            name='SMA 20', 
            line=dict(color='rgb(230, 126, 34)', width=2),
            hovertemplate='<b>Date:</b> %{x}<br><b>SMA 20:</b> ₹%{y:.2f}<extra></extra>'
        ), row=1, col=1)
    
    if 'SMA_50' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['SMA_50'],
            name='SMA 50', 
            line=dict(color='rgb(231, 76, 60)', width=2),
            hovertemplate='<b>Date:</b> %{x}<br><b>SMA 50:</b> ₹%{y:.2f}<extra></extra>'
        ), row=1, col=1)
    
    # Bollinger Bands
    if 'BB_Upper' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['BB_Upper'],
            name='BB Upper', 
            line=dict(color='rgba(155, 89, 182, 0.5)', width=1, dash='dot'),
            showlegend=False
        ), row=1, col=1)
        
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['BB_Lower'],
            name='BB Lower', 
            line=dict(color='rgba(155, 89, 182, 0.5)', width=1, dash='dot'),
            fill='tonexty',
            fillcolor='rgba(155, 89, 182, 0.1)',
            showlegend=False
        ), row=1, col=1)
    
    # RSI & Stochastic
    if 'RSI' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['RSI'],
            name='RSI', 
            line=dict(color='rgb(155, 89, 182)', width=2),
            hovertemplate='<b>Date:</b> %{x}<br><b>RSI:</b> %{y:.2f}<extra></extra>'
        ), row=2, col=1)
        
        fig.add_hline(y=70, line=dict(width=1, dash="dash", color="red"), row=2, col=1)
        fig.add_hline(y=30, line=dict(width=1, dash="dash", color="green"), row=2, col=1)
    
    if '%K' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['%K'],
            name='%K (Stoch)', 
            line=dict(color='rgb(241, 196, 15)', width=1.5),
            hovertemplate='<b>Date:</b> %{x}<br><b>%K:</b> %{y:.2f}<extra></extra>'
        ), row=2, col=1)
    
    # MACD
    if 'MACD' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['MACD'],
            name='MACD', 
            line=dict(color='rgb(52, 152, 219)', width=2),
            hovertemplate='<b>Date:</b> %{x}<br><b>MACD:</b> %{y:.4f}<extra></extra>'
        ), row=3, col=1)
        
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['MACD_Signal'],
            name='Signal', 
            line=dict(color='rgb(231, 76, 60)', width=2),
            hovertemplate='<b>Date:</b> %{x}<br><b>Signal:</b> %{y:.4f}<extra></extra>'
        ), row=3, col=1)
        
        if 'MACD_Histogram' in hist_data.columns:
            colors = ['green' if val >= 0 else 'red' for val in hist_data['MACD_Histogram']]
            fig.add_trace(go.Bar(
                x=hist_data.index, 
                y=hist_data['MACD_Histogram'],
                name='MACD Hist',
                marker_color=colors,
                opacity=0.6,
                showlegend=False
            ), row=3, col=1)
    
    # Volume & ATR
    fig.add_trace(go.Bar(
        x=hist_data.index, 
        y=hist_data['Volume'],
        name='Volume',
        marker_color='rgba(149, 165, 166, 0.7)',
        yaxis='y4',
        hovertemplate='<b>Date:</b> %{x}<br><b>Volume:</b> %{y:,.0f}<extra></extra>'
    ), row=4, col=1)
    
    if 'ATR' in hist_data.columns:
        fig.add_trace(go.Scatter(
            x=hist_data.index, 
            y=hist_data['ATR'],
            name='ATR', 
            line=dict(color='rgb(230, 126, 34)', width=2),
            yaxis='y5',
            hovertemplate='<b>Date:</b> %{x}<br><b>ATR:</b> %{y:.2f}<extra></extra>'
        ), row=4, col=1)
    
    fig.update_layout(
        height=900, 
        showlegend=True,
        plot_bgcolor='rgba(248, 249, 250, 0.8)',
        paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=10),
        hovermode='x unified'
    )
    
    fig.update_yaxes(title_text="Price (₹)", row=1, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(title_text="RSI/Stoch", row=2, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(title_text="MACD", row=3, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_yaxes(title_text="Volume", row=4, col=1, gridcolor='rgba(128, 128, 128, 0.2)')
    fig.update_xaxes(gridcolor='rgba(128, 128, 128, 0.2)')
    
    return fig

# ---------------- ADVANCED ANALYTICS FUNCTIONS ----------------
def calculate_volatility_metrics(df, hist_data):
    """Calculate advanced volatility metrics."""
    if df.empty or hist_data.empty:
        return {}
    
    metrics = {}
    
    # IV Percentiles
    if 'CALL_IV' in df.columns and 'PUT_IV' in df.columns:
        call_iv_percentile = np.percentile(df['CALL_IV'].dropna(), [25, 50, 75])
        put_iv_percentile = np.percentile(df['PUT_IV'].dropna(), [25, 50, 75])
        
        metrics['call_iv_25pct'] = call_iv_percentile[0]
        metrics['call_iv_median'] = call_iv_percentile[1]
        metrics['call_iv_75pct'] = call_iv_percentile[2]
        metrics['put_iv_25pct'] = put_iv_percentile[0]
        metrics['put_iv_median'] = put_iv_percentile[1]
        metrics['put_iv_75pct'] = put_iv_percentile[2]
    
    # IV Rank calculation
    if not hist_data.empty and len(hist_data) > 20:
        # Calculate historical volatility
        returns = hist_data['Close'].pct_change().dropna()
        hist_vol = returns.std() * np.sqrt(252) * 100  # Annualized
        
        # Compare with current IV (using average of call and put IV)
        current_iv = (df['CALL_IV'].mean() + df['PUT_IV'].mean()) / 2 if not df.empty else 20
        
        # Simplified IV Rank calculation
        iv_rank = min(100, max(0, (current_iv - 10) / (50 - 10) * 100))  # Assuming range 10-50%
        metrics['iv_rank'] = iv_rank
    
    return metrics

def calculate_gamma_exposure(df, spot_price):
    """Calculate gamma exposure metrics."""
    if df.empty:
        return {}
    
    metrics = {}
    
    # Simplified gamma exposure calculation
    df_gamma = df.copy()
    df_gamma['Total_Gamma'] = df_gamma['CALL_OI'] + df_gamma['PUT_OI']
    
    if not df_gamma.empty:
        max_gamma_strike = df_gamma.loc[df_gamma['Total_Gamma'].idxmax(), 'STRIKE'] if not df_gamma.empty else 0
        metrics['peak_gamma_strike'] = max_gamma_strike
        
        # Gamma exposure by strike range
        if spot_price > 0:
            atm_range = df_gamma[
                (df_gamma['STRIKE'] >= spot_price * 0.95) & 
                (df_gamma['STRIKE'] <= spot_price * 1.05)
            ]
            if not atm_range.empty:
                atm_gamma = atm_range['Total_Gamma'].sum()
                total_gamma = df_gamma['Total_Gamma'].sum()
                atm_gamma_pct = (atm_gamma / total_gamma * 100) if total_gamma > 0 else 0
                metrics['atm_gamma_concentration'] = atm_gamma_pct
    
    return metrics

def calculate_option_flow(df):
    """Calculate option flow metrics."""
    if df.empty:
        return {}
    
    metrics = {}
    
    # Identify unusual option activity
    if 'CALL_CHNG_IN_OI' in df.columns and 'PUT_CHNG_IN_OI' in df.columns:
        # Find largest OI changes
        max_call_oi_change = df['CALL_CHNG_IN_OI'].max()
        max_put_oi_change = df['PUT_CHNG_IN_OI'].max()
        
        call_oi_strike = df.loc[df['CALL_CHNG_IN_OI'].idxmax(), 'STRIKE'] if max_call_oi_change > 0 else 0
        put_oi_strike = df.loc[df['PUT_CHNG_IN_OI'].idxmax(), 'STRIKE'] if max_put_oi_change > 0 else 0
        
        metrics['largest_call_oi_change'] = max_call_oi_change
        metrics['largest_call_oi_strike'] = call_oi_strike
        metrics['largest_put_oi_change'] = max_put_oi_change
        metrics['largest_put_oi_strike'] = put_oi_strike
    
    # Volume analysis
    if 'CALL_Volume' in df.columns and 'PUT_Volume' in df.columns:
        total_call_volume = df['CALL_Volume'].sum()
        total_put_volume = df['PUT_Volume'].sum()
        volume_ratio = total_put_volume / total_call_volume if total_call_volume > 0 else 0
        metrics['volume_pcr'] = volume_ratio
    
    return metrics

def calculate_adx(hist_data, period=14):
    """Calculate Average Directional Index (ADX)."""
    if hist_data.empty or len(hist_data) < period * 2:
        return pd.DataFrame()
    
    hist_data = hist_data.copy()
    high = hist_data['High']
    low = hist_data['Low']
    close = hist_data['Close']
    
    # +DM and -DM
    plus_dm = high.diff()
    minus_dm = -low.diff()
    
    # True Range
    tr1 = high - low
    tr2 = (high - close.shift()).abs()
    tr3 = (low - close.shift()).abs()
    true_range = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    
    # 14-period averages
    atr = true_range.rolling(period).mean()
    plus_di = 100 * (plus_dm.rolling(period).mean() / atr)
    minus_di = 100 * (minus_dm.rolling(period).mean() / atr)
    
    # ADX calculation
    dx = (abs(plus_di - minus_di) / (plus_di + minus_di)) * 100
    adx = dx.rolling(period).mean()
    
    # Add to dataframe
    hist_data['ADX'] = adx
    hist_data['Plus_DI'] = plus_di
    hist_data['Minus_DI'] = minus_di
    
    return hist_data

def calculate_market_profile(hist_data, lookback_period=20):
    """Calculate market profile metrics."""
    if hist_data.empty or len(hist_data) < lookback_period:
        return {}
    
    recent_data = hist_data.tail(lookback_period)
    
    metrics = {}
    metrics['poc'] = recent_data['Close'].mode().iloc[0] if not recent_data['Close'].mode().empty else recent_data['Close'].iloc[-1]
    metrics['value_area_high'] = recent_data['High'].quantile(0.7)
    metrics['value_area_low'] = recent_data['Low'].quantile(0.3)
    metrics['value_area_width'] = metrics['value_area_high'] - metrics['value_area_low']
    
    # Check if current price is in value area
    current_price = hist_data['Close'].iloc[-1]
    metrics['price_in_value_area'] = metrics['value_area_low'] <= current_price <= metrics['value_area_high']
    
    return metrics

def calculate_seasonality(hist_data):
    """Calculate seasonal patterns."""
    if hist_data.empty or len(hist_data) < 5:
        return {}
    
    metrics = {}
    
    # Analyze performance by day of week
    hist_data['DayOfWeek'] = hist_data.index.dayofweek
    daily_returns = hist_data.groupby('DayOfWeek')['Close'].pct_change().mean()
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    for i, day in enumerate(days):
        if i in daily_returns.index:
            metrics[f'{day.lower()}_avg_return'] = daily_returns[i] * 100
    
    # Best and worst performing days
    if not daily_returns.empty:
        best_day_idx = daily_returns.idxmax()
        worst_day_idx = daily_returns.idxmin()
        metrics['best_day'] = days[best_day_idx] if 0 <= best_day_idx < 5 else "N/A"
        metrics['worst_day'] = days[worst_day_idx] if 0 <= worst_day_idx < 5 else "N/A"
    
    return metrics

# ---------------- EXPORT FUNCTIONALITY ----------------
def export_to_excel(df, analytics, symbol, charts=None):
    """Export data and charts to Excel with professional formatting."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(SAVE_FOLDER, f"{symbol}_OptionChain_{timestamp}.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Option Chain Data"
    
    # Add headers with formatting
    headers = ['Strike', 'Call OI', 'Call Change', 'Call IV', 'Call LTP',
               'Put OI', 'Put Change', 'Put IV', 'Put LTP']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Add analytics summary
    summary_ws = wb.create_sheet("Analytics Summary")
    summary_data = [
        ['Metric', 'Value'],
        ['PCR', analytics.get('pcr', 0)],
        ['ATM PCR', analytics.get('atm_pcr', 0)],
        ['Max Pain', analytics.get('max_pain', 0)],
        ['Support Level', analytics.get('support', 0)],
        ['Resistance Level', analytics.get('resistance', 0)],
        ['Avg Call IV', analytics.get('avg_call_iv', 0)],
        ['Avg Put IV', analytics.get('avg_put_iv', 0)],
        ['IV Skew', analytics.get('iv_skew', 0)],
        ['Expected Move (30D)', analytics.get('expected_move_30d', 0)],
        ['Directional Bias', analytics.get('directional_bias', 'Neutral')]
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary headers
    for cell in summary_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    wb.save(filename)
    return filename

# ---------------- WATCHLIST MANAGEMENT ----------------
def manage_watchlist():
    """Enhanced watchlist management with popular symbols."""
    st.sidebar.markdown("#### 📋 Watchlist Management")
    
    # Quick add popular symbols
    st.sidebar.markdown("##### 🚀 Quick Add Popular:")
    popular_symbols = ["TCS", "RELIANCE", "INFY", "HDFCBANK", "ICICIBANK", "SBIN"]
    
    cols = st.sidebar.columns(2)
    for i, symbol in enumerate(popular_symbols):
        with cols[i % 2]:
            if st.button(f"+ {symbol}", key=f"quick_add_{symbol}"):
                if symbol not in st.session_state.watchlist:
                    st.session_state.watchlist.append(symbol)
                    st.success(f"Added {symbol}")
                    st.rerun()
    
    # Add new symbol
    new_symbol = st.sidebar.text_input("Add Custom Symbol:", placeholder="e.g., BHARTIARTL, LT")
    
    col1, col2 = st.sidebar.columns(2)
    with col1:
        if st.button("➕ Add"):
            if new_symbol and new_symbol.upper() not in st.session_state.watchlist:
                st.session_state.watchlist.append(new_symbol.upper())
                st.success(f"Added {new_symbol.upper()} to watchlist")
                st.rerun()
    
    with col2:
        if st.button("🗑️ Clear All"):
            st.session_state.watchlist = []
            st.success("Watchlist cleared")
            st.rerun()
    
    # Display current watchlist
    if st.session_state.watchlist:
        st.sidebar.markdown("##### Current Watchlist:")
        for symbol in st.session_state.watchlist:
            col1, col2 = st.sidebar.columns([3, 1])
            with col1:
                st.write(f"• {symbol}")
            with col2:
                if st.button("❌", key=f"remove_{symbol}"):
                    st.session_state.watchlist.remove(symbol)
                    st.rerun()
# Add this function to calculate additional technical indicators
def calculate_advanced_indicators(df, period="1d"):
    """Calculate advanced technical indicators including Bollinger Bands, Fibonacci, EMAs, and RSI."""
    if df.empty:
        return df
    
    df = df.copy()
    
    # Calculate EMAs
    df['EMA_20'] = talib.EMA(df['Close'], timeperiod=20)
    df['EMA_50'] = talib.EMA(df['Close'], timeperiod=50)
    df['EMA_75'] = talib.EMA(df['Close'], timeperiod=75)
    
    # Calculate RSI
    df['RSI'] = talib.RSI(df['Close'], timeperiod=14)
    
    # Calculate Bollinger Bands
    df['BB_Upper'], df['BB_Middle'], df['BB_Lower'] = talib.BBANDS(
        df['Close'], timeperiod=20, nbdevup=2, nbdevdn=2, matype=0
    )
    
    # Calculate Fibonacci levels
    if len(df) > 0:
        high = df['High'].max()
        low = df['Low'].min()
        diff = high - low
        
        df['Fib_0'] = high
        df['Fib_0.236'] = high - 0.236 * diff
        df['Fib_0.382'] = high - 0.382 * diff
        df['Fib_0.5'] = high - 0.5 * diff
        df['Fib_0.618'] = high - 0.618 * diff
        df['Fib_0.786'] = high - 0.786 * diff
        df['Fib_1'] = low
    
    # Calculate MACD
    df['MACD'], df['MACD_Signal'], df['MACD_Hist'] = talib.MACD(
        df['Close'], fastperiod=12, slowperiod=26, signalperiod=9
    )
    
    # Calculate Stochastic
    df['STOCH_K'], df['STOCH_D'] = talib.STOCH(
        df['High'], df['Low'], df['Close'], 
        fastk_period=14, slowk_period=3, slowk_matype=0, slowd_period=3, slowd_matype=0
    )
    
    # Generate buy/sell signals
    df = generate_signals(df)
    
    return df

def generate_signals(df):
    """Generate buy and sell signals based on multiple indicators."""
    if df.empty:
        return df
    
    df = df.copy()
    
    # Initialize signal columns
    df['Signal'] = 0  # 0: No signal, 1: Buy, -1: Sell
    df['Signal_Strength'] = 0
    
    # RSI signals
    df.loc[df['RSI'] < 30, 'Signal'] = 1  # Oversold - Buy
    df.loc[df['RSI'] > 70, 'Signal'] = -1  # Overbought - Sell
    
    # EMA crossover signals
    df['EMA_20_50_Cross'] = np.where(df['EMA_20'] > df['EMA_50'], 1, -1)
    df.loc[df['EMA_20_50_Cross'] == 1, 'Signal'] = 1
    df.loc[df['EMA_20_50_Cross'] == -1, 'Signal'] = -1
    
    # Bollinger Band signals
    df.loc[df['Close'] < df['BB_Lower'], 'Signal'] = 1  # Price below lower band - Buy
    df.loc[df['Close'] > df['BB_Upper'], 'Signal'] = -1  # Price above upper band - Sell
    
    # MACD signals
    df['MACD_Cross'] = np.where(df['MACD'] > df['MACD_Signal'], 1, -1)
    df.loc[df['MACD_Cross'] == 1, 'Signal'] = 1
    df.loc[df['MACD_Cross'] == -1, 'Signal'] = -1
    
    # Stochastic signals
    df.loc[(df['STOCH_K'] < 20) & (df['STOCH_D'] < 20), 'Signal'] = 1  # Oversold - Buy
    df.loc[(df['STOCH_K'] > 80) & (df['STOCH_D'] > 80), 'Signal'] = -1  # Overbought - Sell
    
    # Calculate signal strength based on confluence
    for index, row in df.iterrows():
        strength = 0
        
        # RSI contribution
        if row['RSI'] < 30:
            strength += 1
        elif row['RSI'] > 70:
            strength -= 1
        
        # EMA crossover contribution
        if row['EMA_20'] > row['EMA_50']:
            strength += 1
        else:
            strength -= 1
        
        # Bollinger Band contribution
        if row['Close'] < row['BB_Lower']:
            strength += 1
        elif row['Close'] > row['BB_Upper']:
            strength -= 1
        
        # MACD contribution
        if row['MACD'] > row['MACD_Signal']:
            strength += 1
        else:
            strength -= 1
        
        # Stochastic contribution
        if row['STOCH_K'] < 20 and row['STOCH_D'] < 20:
            strength += 1
        elif row['STOCH_K'] > 80 and row['STOCH_D'] > 80:
            strength -= 1
        
        df.at[index, 'Signal_Strength'] = strength
    
    return df

# Add this function to create the intraday analysis section
def create_intraday_analysis_tab(hist_data, symbol):
    """Create a comprehensive intraday analysis tab with multiple timeframes and signal analysis."""
    if hist_data.empty:
        return st.warning("No historical data available for intraday analysis.")
    
    st.markdown(f"### 📊 Intraday Analysis for {symbol}")
    
    # Timeframe selection
    timeframe = st.selectbox(
        "Select Timeframe:",
        ["5 min", "15 min", "30 min", "120 min", "240 min", "Day", "Week", "Month"],
        index=2  # Default to 30min
    )
    
    # Resample data based on selected timeframe
    if timeframe == "5 min":
        resample_period = "5T"
    elif timeframe == "15 min":
        resample_period = "15T"
    elif timeframe == "30 min":
        resample_period = "30T"
    elif timeframe == "120 min":
        resample_period = "2H"
    elif timeframe == "240 min":
        resample_period = "4H"
    elif timeframe == "Day":
        resample_period = "1D"
    elif timeframe == "Week":
        resample_period = "1W"
    else:  # Month
        resample_period = "1M"
    
    # Resample the data
    intraday_data = hist_data.resample(resample_period).agg({
        'Open': 'first',
        'High': 'max',
        'Low': 'min',
        'Close': 'last',
        'Volume': 'sum'
    }).dropna()
    
    # Calculate indicators for the resampled data
    intraday_data = calculate_advanced_indicators(intraday_data)
    
    # Display current signals
    if not intraday_data.empty:
        latest_signal = intraday_data['Signal'].iloc[-1]
        signal_strength = intraday_data['Signal_Strength'].iloc[-1]
        current_price = intraday_data['Close'].iloc[-1]
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Current Price", f"₹{current_price:.2f}")
        
        with col2:
            if latest_signal == 1:
                st.metric("Signal", "BUY", delta="Bullish", delta_color="normal")
            elif latest_signal == -1:
                st.metric("Signal", "SELL", delta="Bearish", delta_color="inverse")
            else:
                st.metric("Signal", "NEUTRAL")
        
        with col3:
            st.metric("Signal Strength", f"{abs(signal_strength)}/5", 
                     delta="Strong" if abs(signal_strength) >= 3 else "Moderate" if abs(signal_strength) >= 2 else "Weak")
        
        with col4:
            # Calculate signal accuracy if we have enough historical data
            if len(intraday_data) > 10:
                profitable_signals = 0
                total_signals = 0
                
                for i in range(1, len(intraday_data)):
                    if intraday_data['Signal'].iloc[i-1] != 0:  # Had a signal
                        total_signals += 1
                        price_change = intraday_data['Close'].iloc[i] - intraday_data['Close'].iloc[i-1]
                        
                        # Check if signal was correct
                        if (intraday_data['Signal'].iloc[i-1] == 1 and price_change > 0) or \
                           (intraday_data['Signal'].iloc[i-1] == -1 and price_change < 0):
                            profitable_signals += 1
                
                accuracy = (profitable_signals / total_signals * 100) if total_signals > 0 else 0
                st.metric("Signal Accuracy", f"{accuracy:.1f}%")
    
    # Create the chart
    fig = make_subplots(
        rows=3, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.05,
        subplot_titles=(
            f'Price with Indicators ({timeframe})',
            'RSI and Stochastic',
            'MACD'
        ),
        row_heights=[0.5, 0.25, 0.25]
    )
    
    # Price chart with indicators
    fig.add_trace(go.Candlestick(
        x=intraday_data.index,
        open=intraday_data['Open'],
        high=intraday_data['High'],
        low=intraday_data['Low'],
        close=intraday_data['Close'],
        name='Price'
    ), row=1, col=1)
    
    # Add EMAs
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['EMA_20'],
        name='EMA 20', line=dict(color='orange', width=1)
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['EMA_50'],
        name='EMA 50', line=dict(color='blue', width=1)
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['EMA_75'],
        name='EMA 75', line=dict(color='purple', width=1)
    ), row=1, col=1)
    
    # Add Bollinger Bands
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['BB_Upper'],
        name='BB Upper', line=dict(color='gray', width=1, dash='dash'),
        opacity=0.7
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['BB_Middle'],
        name='BB Middle', line=dict(color='gray', width=1),
        opacity=0.7
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['BB_Lower'],
        name='BB Lower', line=dict(color='gray', width=1, dash='dash'),
        opacity=0.7,
        fill='tonexty', fillcolor='rgba(128,128,128,0.1)'
    ), row=1, col=1)
    
    # Add buy/sell signals
    buy_signals = intraday_data[intraday_data['Signal'] == 1]
    sell_signals = intraday_data[intraday_data['Signal'] == -1]
    
    if not buy_signals.empty:
        fig.add_trace(go.Scatter(
            x=buy_signals.index, y=buy_signals['Low'] * 0.995,
            mode='markers', name='Buy Signal',
            marker=dict(color='green', size=10, symbol='triangle-up'),
            hovertemplate='<b>Buy Signal</b><br>Strength: %{text}<extra></extra>',
            text=[f"{abs(s):.0f}/5" for s in buy_signals['Signal_Strength']]
        ), row=1, col=1)
    
    if not sell_signals.empty:
        fig.add_trace(go.Scatter(
            x=sell_signals.index, y=sell_signals['High'] * 1.005,
            mode='markers', name='Sell Signal',
            marker=dict(color='red', size=10, symbol='triangle-down'),
            hovertemplate='<b>Sell Signal</b><br>Strength: %{text}<extra></extra>',
            text=[f"{abs(s):.0f}/5" for s in sell_signals['Signal_Strength']]
        ), row=1, col=1)
    
    # RSI and Stochastic
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['RSI'],
        name='RSI', line=dict(color='blue', width=1)
    ), row=2, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['STOCH_K'],
        name='Stoch %K', line=dict(color='orange', width=1)
    ), row=2, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['STOCH_D'],
        name='Stoch %D', line=dict(color='green', width=1)
    ), row=2, col=1)
    
    # Add RSI levels
    fig.add_hline(y=70, line_dash="dash", line_color="red", row=2, col=1)
    fig.add_hline(y=30, line_dash="dash", line_color="green", row=2, col=1)
    fig.add_hline(y=50, line_dash="dot", line_color="gray", row=2, col=1)
    
    # MACD
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['MACD'],
        name='MACD', line=dict(color='blue', width=1)
    ), row=3, col=1)
    
    fig.add_trace(go.Scatter(
        x=intraday_data.index, y=intraday_data['MACD_Signal'],
        name='Signal', line=dict(color='red', width=1)
    ), row=3, col=1)
    
    # MACD histogram
    colors = np.where(intraday_data['MACD_Hist'] < 0, 'red', 'green')
    fig.add_trace(go.Bar(
        x=intraday_data.index, y=intraday_data['MACD_Hist'],
        name='Histogram', marker_color=colors,
        opacity=0.5
    ), row=3, col=1)
    
    # Add zero line for MACD
    fig.add_hline(y=0, line_dash="solid", line_color="black", row=3, col=1)
    
    # Update layout
    fig.update_layout(
        height=800,
        title=f"Intraday Analysis - {symbol} ({timeframe})",
        xaxis_rangeslider_visible=False,
        showlegend=True
    )
    
    # Update y-axis labels
    fig.update_yaxes(title_text="Price", row=1, col=1)
    fig.update_yaxes(title_text="RSI/Stoch", row=2, col=1)
    fig.update_yaxes(title_text="MACD", row=3, col=1)
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Display signal details
    st.markdown("##### 📋 Signal Details")
    
    if not intraday_data.empty:
        latest = intraday_data.iloc[-1]
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("RSI", f"{latest['RSI']:.2f}")
        
        with col2:
            st.metric("Stoch %K", f"{latest['STOCH_K']:.2f}")
        
        with col3:
            st.metric("MACD", f"{latest['MACD']:.4f}")
        
        with col4:
            ema_trend = "Bullish" if latest['Close'] > latest['EMA_20'] else "Bearish"
            st.metric("EMA Trend", ema_trend)
    
    # Signal performance analysis
    st.markdown("##### 📊 Signal Performance Analysis")
    
    if len(intraday_data) > 10:
        # Calculate signal performance metrics
        buy_signals = intraday_data[intraday_data['Signal'] == 1]
        sell_signals = intraday_data[intraday_data['Signal'] == -1]
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Buy Signals", len(buy_signals))
        
        with col2:
            st.metric("Total Sell Signals", len(sell_signals))
        
        with col3:
            if len(buy_signals) > 0:
                avg_buy_strength = buy_signals['Signal_Strength'].abs().mean()
                st.metric("Avg Buy Strength", f"{avg_buy_strength:.1f}/5")
            else:
                st.metric("Avg Buy Strength", "N/A")
        
        with col4:
            if len(sell_signals) > 0:
                avg_sell_strength = sell_signals['Signal_Strength'].abs().mean()
                st.metric("Avg Sell Strength", f"{avg_sell_strength:.1f}/5")
            else:
                st.metric("Avg Sell Strength", "N/A")
    
    # Show recent signals table
    st.markdown("##### 📈 Recent Signals")
    
    recent_signals = intraday_data[intraday_data['Signal'] != 0].tail(10)
    
    if not recent_signals.empty:
        signal_df = recent_signals[['Close', 'Signal', 'Signal_Strength', 'RSI', 'STOCH_K', 'MACD']].copy()
        signal_df['Signal'] = signal_df['Signal'].apply(lambda x: 'BUY' if x == 1 else 'SELL')
        signal_df['Time'] = signal_df.index
        signal_df = signal_df[['Time', 'Close', 'Signal', 'Signal_Strength', 'RSI', 'STOCH_K', 'MACD']]
        
        # Format the table with colors
        def color_signal(val):
            if val == 'BUY':
                return 'background-color: lightgreen'
            elif val == 'SELL':
                return 'background-color: lightcoral'
            return ''
        
        styled_df = signal_df.style.apply(
            lambda x: ['background: lightgreen' if x.Signal == 'BUY' else 'background: lightcoral' for _ in x], 
            axis=1
        ).format({
            'Close': '₹{:.2f}',
            'Signal_Strength': '{:.0f}',
            'RSI': '{:.2f}',
            'STOCH_K': '{:.2f}',
            'MACD': '{:.4f}'
        })
        
        st.dataframe(styled_df, width='stretch')
        
        # Add signal statistics
        if len(recent_signals) >= 3:
            st.markdown("##### 📈 Signal Statistics")
            
            # Calculate win rate for recent signals
            wins = 0
            total_trades = 0
            
            for i in range(len(recent_signals)):
                if i < len(recent_signals) - 1:
                    signal = recent_signals['Signal'].iloc[i]
                    entry_price = recent_signals['Close'].iloc[i]
                    exit_price = recent_signals['Close'].iloc[i+1]
                    
                    if signal == 1:  # Buy signal
                        if exit_price > entry_price:
                            wins += 1
                        total_trades += 1
                    elif signal == -1:  # Sell signal
                        if exit_price < entry_price:
                            wins += 1
                        total_trades += 1
            
            if total_trades > 0:
                win_rate = (wins / total_trades) * 100
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Recent Win Rate", f"{win_rate:.1f}%")
                with col2:
                    st.metric("Recent Trades", total_trades)
                with col3:
                    st.metric("Recent Wins", wins)
    else:
        st.info("No recent signals generated.")
# Add this import statement with your other imports
try:
    from advanced_analysis import (
        create_sector_analysis_tab, 
        create_market_sentiment_tab,
        create_decision_dashboard,
        create_advanced_analytics_tab
    )
except ImportError:
    st.warning("Advanced analysis modules not available")
# Import the new modules
try:
    from live_dashboard import create_live_dashboard_tab
    from news_social import create_news_social_tab
    from excel_automation import create_excel_automation_tab
except ImportError:
    st.warning("Some advanced modules could not be loaded. Some features may be limited.")
# ---------------- MAIN APP ----------------
def main():
    """Enhanced main Streamlit application."""
    st.set_page_config(
        page_title="NSE Option Chain Analytics Pro",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # Apply custom CSS
    st.markdown(PRO_CSS, unsafe_allow_html=True)

    # Header
    st.markdown(
        """
    <div class="main-header">
        <h1>🚀 NSE Option Chain Analytics Pro</h1>
        <p>Advanced Market Analysis with Live Dashboard & Automated Reporting</p>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # Sidebar configuration
    st.sidebar.markdown("### ⚙️ Configuration")

    # Symbol type selection
    symbol_type = st.sidebar.radio(
        "Select Data Type:", ["Indices", "Popular Stocks", "IT Sector", "Banking", "Custom"]
    )

    # Symbol selection
    if symbol_type == "Indices":
        available_symbols = list(st.session_state.indices) + [
            "NIFTYIT", "NIFTYPHARMA", "NIFTYAUTO", "NIFTYMETAL"
        ]
    elif symbol_type == "Popular Stocks":
        available_symbols = ["RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","ITC","LT","BHARTIARTL","ASIANPAINT","MARUTI"]
    elif symbol_type == "IT Sector":
        available_symbols = ["TCS","INFY","WIPRO","HCLTECH","TECHM","LTIM","MINDTREE","COFORGE"]
    elif symbol_type == "Banking":
        available_symbols = ["HDFCBANK","ICICIBANK","SBIN","KOTAKBANK","AXISBANK","INDUSINDBK","BANDHANBNK"]
    else:  # Custom
        available_symbols = st.session_state.watchlist if st.session_state.watchlist else ["TCS"]

    symbol = st.sidebar.selectbox(
        "Select Symbol:", options=available_symbols, index=0 if available_symbols else None
    )

    # Manual input for custom
    if symbol_type == "Custom":
        manual_symbol = st.sidebar.text_input("Or enter symbol manually:", placeholder="e.g., BHARTIARTL, NESTLEIND")
        if manual_symbol:
            symbol = manual_symbol.upper()

    # Data source & historical period
    st.sidebar.markdown("#### 📡 Data Source Options")
    use_alternative_source = st.sidebar.checkbox("Use Alternative NSE API", value=False)
    historical_period = st.sidebar.selectbox(
        "Historical Data Period", ["1d","5d","1mo","3mo","6mo","1y","2y"], index=4
    )

    # Auto-refresh
    auto_refresh = st.sidebar.checkbox("🔄 Auto Refresh", value=st.session_state.auto_refresh)
    refresh_interval = st.sidebar.slider("Refresh Interval (seconds)", 10, 300, 30)
    st.session_state.auto_refresh = auto_refresh

    # Watchlist management
    manage_watchlist()

    # Risk settings
    st.sidebar.markdown("#### 🎯 Risk Parameters")
    risk_free_rate = st.sidebar.slider("Risk Free Rate (%)", 0.0, 10.0, 5.0)/100
    days_to_expiry = st.sidebar.number_input("Days to Expiry", 1, 365, 30)

    # Fetch data
    with st.spinner(f"Fetching {symbol} data..."):
        hist_data = fetch_historical_data(symbol, period=historical_period)

        if hist_data.empty:
            st.error(f"❌ Unable to fetch historical data for {symbol}.")
            st.info("💡 Try another symbol or period")
            return
        else:
            hist_data = calculate_technical_indicators(hist_data)
            current_price = hist_data["Close"].iloc[-1]
            price_change = hist_data["Close"].iloc[-1] - hist_data["Close"].iloc[-2] if len(hist_data)>1 else 0
            price_change_pct = (price_change/hist_data["Close"].iloc[-2]*100) if len(hist_data)>1 and hist_data["Close"].iloc[-2]!=0 else 0
            st.info(f"📈 **{symbol}** - Current Price: ₹{current_price:.2f} ({price_change:+.2f}, {price_change_pct:+.2f}%)")

        # Option chain
        if use_alternative_source:
            df_alt = fetch_nse_options_alternative(symbol)
            if not df_alt.empty:
                df = df_alt
                st.success("✅ Using alternative NSE data source")
            else:
                session = get_nse_session()
                raw_data = fetch_option_chain(symbol, session)
                df = parse_data(symbol, raw_data)
        else:
            session = get_nse_session()
            raw_data = fetch_option_chain(symbol, session)
            df = parse_data(symbol, raw_data)

    # Analytics
    analytics = calculate_analytics(df, current_price)
    spot_price = analytics.get("spot_price", current_price)

    # Main metrics
    col1,col2,col3,col4,col5 = st.columns(5)
    with col1: st.metric("Current Price", f"₹{current_price:.2f}", delta=f"{price_change:+.2f}")
    with col2: 
        pcr = analytics.get("pcr",0)
        pcr_delta = "📈" if pcr>1.2 else "📉" if pcr<0.8 else "➡️"
        st.metric("PCR", f"{pcr:.2f} {pcr_delta}" if pcr>0 else "N/A")
    with col3: st.metric("Max Pain", f"₹{analytics.get('max_pain',0):,.0f}" if analytics.get("max_pain",0)>0 else "N/A")
    with col4: st.metric("IV ATM", f"{analytics.get('iv_atm',0):.2f}%" if analytics.get("iv_atm",0)>0 else "N/A")
    with col5:
        bias = analytics.get("directional_bias","Neutral")
        bias_emoji = "🟢" if bias=="Bullish" else "🔴" if bias=="Bearish" else "🟡"
        st.metric("Bias", f"{bias} {bias_emoji}")

    # ---------------- Primary Tabs ----------------
    primary_tabs = st.tabs([
        "📊 Option Chain","📈 Charts","🧠 ML Predictions","📋 Technical Analysis",
        "🎯 Strategy Builder","💾 Export","🔍 More Analysis","📱 Intraday Analysis"
    ])

    with primary_tabs[0]:
        st.markdown("### 📊 Option Chain Data")
        if df.empty:
            st.warning("⚠️ Option chain data not available")
        else:
            # Display key metrics, support/resistance, option flow
            col1, col2, col3 = st.columns(3)
            with col1:
                metrics_data = {"Metric":["PCR","ATM PCR","IV Skew","Expected Move"],
                                "Value":[f"{analytics.get('pcr',0):.2f}",
                                         f"{analytics.get('atm_pcr',0):.2f}",
                                         f"{analytics.get('iv_skew',0):.2f}%",
                                         f"₹{analytics.get('expected_move_30d',0):.2f}"]}
                st.dataframe(create_safe_dataframe(metrics_data), use_container_width=True)
            with col2:
                sr_data = {"Level":["Strong Support","Max Pain","Strong Resistance"],
                           "Strike":[f"₹{analytics.get('support',0):,.0f}",
                                     f"₹{analytics.get('max_pain',0):,.0f}",
                                     f"₹{analytics.get('resistance',0):,.0f}"]}
                st.dataframe(create_safe_dataframe(sr_data), use_container_width=True)
            with col3:
                flow_data = {"Type":["Call Flow","Put Flow","Net Flow"],
                             "Value":[f"₹{analytics.get('call_flow',0):,.2f}",
                                      f"₹{analytics.get('put_flow',0):,.2f}",
                                      f"₹{analytics.get('net_flow',0):,.2f}"]}
                st.dataframe(create_safe_dataframe(flow_data), use_container_width=True)
            # Option chain table
            display_df = df[["STRIKE","CALL_OI","CALL_CHNG_IN_OI","CALL_IV","CALL_LTP",
                             "PUT_LTP","PUT_IV","PUT_CHNG_IN_OI","PUT_OI"]].copy()
            st.dataframe(ensure_dataframe_types(display_df), use_container_width=True, height=400)

    with primary_tabs[1]:
        st.markdown("### 📈 Visual Analytics")
        if not df.empty:
            st.plotly_chart(plot_oi_distribution(df, spot_price), use_container_width=True)
            col1,col2 = st.columns(2)
            with col1: st.plotly_chart(plot_oi_change(df, spot_price), use_container_width=True)
            with col2: st.plotly_chart(plot_iv_surface(df, spot_price), use_container_width=True)
        if not hist_data.empty: st.plotly_chart(plot_technical_indicators(hist_data), use_container_width=True)

    with primary_tabs[2]:
        st.markdown("### 🧠 ML Predictions")
        predictions = predict_price_movement_enhanced(hist_data, analytics, df)
        if predictions:
            num_models = len(predictions)
            cols = st.columns(min(num_models,4))
            for i,(model_name,result) in enumerate(predictions.items()):
                col_idx = i%4
                with cols[col_idx]:
                    color = "green" if result["prediction"]==1 else "red"
                    st.markdown(f"<div class='metric-card'><h4>{model_name}</h4><p style='color:{color}; font-weight:bold;'>{'Bullish 📈' if result['prediction']==1 else 'Bearish 📉'}</p><p>Accuracy: {result['accuracy']:.2%}</p><p>Confidence: {result.get('confidence',0.5):.2%}</p></div>",unsafe_allow_html=True)
            model_df = create_safe_dataframe({"Model":list(predictions.keys()),
                                             "Accuracy":[f"{pred['accuracy']:.2%}" for pred in predictions.values()],
                                             "Prediction":["Bullish 📈" if pred["prediction"]==1 else "Bearish 📉" for pred in predictions.values()],
                                             "Confidence":[f"{pred.get('confidence',0.5):.2%}" for pred in predictions.values()]})
            st.dataframe(model_df, use_container_width=True)
        else: st.warning("⚠️ Unable to generate ML predictions.")

    with primary_tabs[3]:
        st.markdown("### 📋 Enhanced Technical Analysis")
        if not hist_data.empty and len(hist_data)>5:
            latest_close = hist_data["Close"].iloc[-1]
            st.metric("Current Price", f"₹{latest_close:.2f}")
        else:
            st.warning("⚠️ Insufficient historical data")

    with primary_tabs[4]:
        st.markdown("### 🎯 Strategy Builder")
        col1,col2 = st.columns([1,1])
        with col1:
            strategy = st.selectbox("Strategy Type",["Long Call","Long Put","Short Call","Short Put","Bull Call Spread","Bear Put Spread","Straddle","Strangle"])
            strike_price = st.number_input("Strike Price", value=int(spot_price), step=50)
            premium = st.number_input("Premium", value=100.0, step=10.0)
            quantity = st.number_input("Quantity", value=1, step=1)
        with col2:
            try:
                greeks = option_greeks(strike=strike_price,premium=premium,spot_price=spot_price,
                                       iv=analytics.get("iv_atm",20),risk_free_rate=risk_free_rate,days_to_expiry=days_to_expiry)
                summary_df = pd.DataFrame.from_dict(strategy_summary(strategy,strike_price,premium,quantity,greeks),orient="index",columns=["Value"])
                st.table(ensure_dataframe_types(summary_df))
                st.plotly_chart(plot_strategy(strategy,strike_price,premium,quantity,spot_price), use_container_width=True)
            except Exception as e:
                st.error(f"Error: {e}")

    with primary_tabs[5]:
        st.markdown("### 💾 Export Data")
        col1,col2 = st.columns(2)
        with col1:
            export_format = st.radio("Export Format:",["Excel","CSV"])
            include_charts = st.checkbox("Include Charts (Excel only)", value=True)
            include_historical = st.checkbox("Include Historical Data", value=True)
            if st.button("💾 Export Data"):
                with st.spinner("Preparing export file..."):
                    try:
                        if export_format=="Excel":
                            charts={}
                            if include_charts and not df.empty:
                                charts = {"OI Distribution":plot_oi_distribution(df,spot_price),
                                          "OI Change":plot_oi_change(df,spot_price),
                                          "IV Surface":plot_iv_surface(df,spot_price)}
                                if not hist_data.empty:
                                    charts["Technical Indicators"]=plot_technical_indicators(hist_data)
                            filename = export_to_excel(df,analytics,symbol,charts)
                            with open(filename,"rb") as f: bytes_data=f.read()
                            st.download_button("📥 Download Excel File", data=bytes_data, file_name=os.path.basename(filename), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        else:
                            csv_data = df.to_csv(index=False) if not df.empty else hist_data.to_csv(index=True) if not hist_data.empty else "No data available"
                            st.download_button("📥 Download CSV File", data=csv_data, file_name=f"{symbol}_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", mime="text/csv")
                    except Exception as e: st.error(f"❌ Export failed: {e}")
        with col2:
            st.info("Excel: Option chain, historical data, analytics, charts\nCSV: Raw data only")

    with primary_tabs[6]:
        st.markdown("### 🔍 Advanced Analytics")
        col1,col2 = st.columns([1,1])
        with col1: 
            if not df.empty:
                volatility_metrics = calculate_volatility_metrics(df,hist_data)
                if volatility_metrics:
                    iv_df = create_safe_dataframe({"Metric":["25th Percentile","Median","75th Percentile"],
                                                   "Call IV":[f"{volatility_metrics.get('call_iv_25pct',0):.2f}%",f"{volatility_metrics.get('call_iv_median',0):.2f}%",f"{volatility_metrics.get('call_iv_75pct',0):.2f}%"],
                                                   "Put IV":[f"{volatility_metrics.get('put_iv_25pct',0):.2f}%",f"{volatility_metrics.get('put_iv_median',0):.2f}%",f"{volatility_metrics.get('put_iv_75pct',0):.2f}%"]})
                    st.dataframe(iv_df,use_container_width=True)
        with col2:
            if not df.empty:
                flow_metrics = calculate_option_flow(df)
                if flow_metrics:
                    flow_df = create_safe_dataframe({"Flow Type":["Largest Call OI Increase","Largest Put OI Increase"],
                                                    "Strike":[f"₹{flow_metrics.get('largest_call_oi_strike',0):,.0f}",f"₹{flow_metrics.get('largest_put_oi_strike',0):,.0f}"],
                                                    "Change":[f"{flow_metrics.get('largest_call_oi_change',0):,.0f}",f"{flow_metrics.get('largest_put_oi_change',0):,.0f}"]})
                    st.dataframe(flow_df,use_container_width=True)
                    st.metric("Volume PCR",f"{flow_metrics.get('volume_pcr',0):.2f}")

    with primary_tabs[7]:
        create_intraday_analysis_tab(hist_data, symbol)

    # ---------------- Secondary Tabs ----------------
    secondary_tabs = st.tabs([
        "🏢 Sector Analysis","📊 Market Sentiment","🎯 Decision Dashboard",
        "📈 Live Dashboard","📰 News & Social","⏰ Excel Automation"
    ])

    with secondary_tabs[0]:
        create_sector_analysis_tab(hist_data, symbol)
    with secondary_tabs[1]:
        create_market_sentiment_tab()
    with secondary_tabs[2]:
        create_decision_dashboard(symbol,hist_data,analytics,predictions)
        with st.expander("Advanced Statistical Analysis"):
            create_advanced_analytics_tab(hist_data, symbol)
    with secondary_tabs[3]:
        create_live_dashboard_tab()
    with secondary_tabs[4]:
        create_news_social_tab()
    with secondary_tabs[5]:
        create_excel_automation_tab()

    # Footer
    st.markdown("---")
    col1,col2,col3,col4 = st.columns(4)
    with col1: st.markdown(f"**Last Updated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    with col2: st.markdown("**Data Sources:** NSE India, Yahoo Finance")
    with col3: st.markdown(f"**Data Points:** {len(hist_data) if not hist_data.empty else 'N/A'}")
    with col4:
        if st.button("🔄 Refresh Data"): st.rerun()

    # Auto-refresh
    if auto_refresh:
        time.sleep(refresh_interval)
        st.rerun()


if __name__=="__main__":
    main()

